"""
Importador genérico CSV/XLSX (Onda 1.6b).

Caso de uso: o operador exportou da prefeitura uma planilha com
matrícula/CPF + alguns campos pré-eSocial (raça, nome da mãe, CEP,
PIS, etc.) e quer aplicar em lote sem editar 200 cadastros à mão.

Padrão de chamada:

    resultado = importar_servidores_csv(
        conteudo_bytes=file.read(),
        nome_arquivo="planilha.csv",
        coluna_identificador="matricula",  # ou "cpf"
        dry_run=False,
    )

Retorno:

    {
        "total_linhas": 200,
        "atualizados": 187,
        "ignorados_servidor_nao_encontrado": 8,
        "ignorados_sem_mudanca": 5,
        "erros": [{"linha": 12, "mensagem": "..."}, ...],
        "preview": [{"linha": 1, "antes": {...}, "depois": {...}}, ...],
    }

Em `dry_run=True`, nada é persistido — só calcula o preview e os
erros previsíveis. Útil pra o operador conferir antes de aplicar.

Não usa pandas/numpy — `csv` da stdlib + `openpyxl` (já dep do Django
admin export) basta para os volumes esperados (até 10k linhas).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from dataclasses import dataclass, field
from typing import Any

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction

from apps.core.validators import validar_cpf
from apps.people.models import Servidor

# Whitelist de colunas aceitas. Chaves são nomes "amigáveis" (case-insensitive,
# espaços/underscores normalizados); valor é o nome do campo do modelo.
# Manter restrito aos campos seguros pra import em massa — qualquer
# campo sensível (cpf, data_nascimento, matricula) fica de fora.
COLUNAS_ACEITAS: dict[str, str] = {
    "tipo_logradouro": "tipo_logradouro",
    "tipo de logradouro": "tipo_logradouro",
    "logradouro": "logradouro",
    "rua": "logradouro",
    "endereco": "logradouro",
    "endereço": "logradouro",
    "numero": "numero",
    "número": "numero",
    "complemento": "complemento",
    "bairro": "bairro",
    "cidade": "cidade",
    "municipio": "cidade",
    "município": "cidade",
    "uf": "uf",
    "estado": "uf",
    "cep": "cep",
    "nacionalidade": "nacionalidade",
    "raca": "raca",
    "raça": "raca",
    "raca cor": "raca",
    "raça cor": "raca",
    "estado civil": "estado_civil",
    "instrucao": "instrucao",
    "instrução": "instrucao",
    "grau de instrucao": "instrucao",
    "grau de instrução": "instrucao",
    "nome do pai": "nome_pai",
    "nome pai": "nome_pai",
    "nome da mae": "nome_mae",
    "nome da mãe": "nome_mae",
    "nome mae": "nome_mae",
    "nome mãe": "nome_mae",
    "pis": "pis_pasep",
    "pis pasep": "pis_pasep",
    "pis_pasep": "pis_pasep",
    "email": "email",
    "telefone": "telefone",
}


@dataclass
class _LinhaResultado:
    linha: int
    identificador: str
    servidor_id: int | None
    antes: dict[str, Any] = field(default_factory=dict)
    depois: dict[str, Any] = field(default_factory=dict)
    mensagem_erro: str = ""

    @property
    def mudou(self) -> bool:
        return self.servidor_id is not None and self.antes != self.depois


def _normalizar_cabecalho(header: str) -> str:
    return " ".join(header.lower().replace("_", " ").split())


def _ler_csv(conteudo: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Decodifica e parseia um CSV em UTF-8 (com fallback BOM)."""
    texto = conteudo.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(texto), delimiter=_detectar_separador(texto))
    rows = list(reader)
    if not rows:
        raise ValueError("Arquivo CSV vazio.")
    cabecalho_bruto = rows[0]
    dados = [
        dict(zip(cabecalho_bruto, linha, strict=False))
        for linha in rows[1:]
        if any(c.strip() for c in linha)
    ]
    return cabecalho_bruto, dados


def _detectar_separador(texto: str) -> str:
    """Heurística simples — comma ou semicolon (Excel BR usa ;)."""
    primeira_linha = texto.splitlines()[0] if texto else ""
    return ";" if primeira_linha.count(";") > primeira_linha.count(",") else ","


def _ler_xlsx(conteudo: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Decodifica um XLSX usando openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Suporte a XLSX requer openpyxl instalado."
        ) from exc

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Planilha vazia.")
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        raise ValueError("Planilha vazia.")
    cabecalho_bruto = [str(c) if c is not None else "" for c in linhas[0]]
    dados: list[dict[str, str]] = []
    for linha in linhas[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in linha):
            continue
        valores = [str(c).strip() if c is not None else "" for c in linha]
        dados.append(dict(zip(cabecalho_bruto, valores, strict=False)))
    return cabecalho_bruto, dados


def _parsear_arquivo(conteudo: bytes, nome_arquivo: str) -> tuple[list[str], list[dict[str, str]]]:
    nome = nome_arquivo.lower()
    if nome.endswith(".xlsx") or nome.endswith(".xlsm"):
        return _ler_xlsx(conteudo)
    return _ler_csv(conteudo)


def _mapear_colunas(cabecalho: Iterable[str]) -> dict[str, str]:
    """
    Para cada coluna no cabeçalho, descobre qual campo do modelo
    ela representa. Colunas desconhecidas são ignoradas (sem erro).
    """
    mapeamento: dict[str, str] = {}
    for col in cabecalho:
        chave = _normalizar_cabecalho(col)
        campo_modelo = COLUNAS_ACEITAS.get(chave)
        if campo_modelo:
            mapeamento[col] = campo_modelo
    return mapeamento


def _normalizar_valor(campo: str, valor: str) -> str:
    valor = (valor or "").strip()
    if campo == "uf":
        return valor.upper()[:2]
    if campo == "cep":
        digitos = "".join(c for c in valor if c.isdigit())
        if len(digitos) == 8:
            return f"{digitos[:5]}-{digitos[5:]}"
        return valor
    return valor


def _resolver_servidor(identificador: str, coluna_identificador: str) -> Servidor | None:
    valor = identificador.strip()
    if not valor:
        return None
    if coluna_identificador == "cpf":
        try:
            cpf_normalizado = validar_cpf(valor)
        except DjangoValidationError:
            return None
        return Servidor.objects.filter(cpf=cpf_normalizado).first()
    return Servidor.objects.filter(matricula=valor).first()


def importar_servidores_csv(  # noqa: C901 — parsing + validação + preview num só lugar é proposital
    *,
    conteudo_bytes: bytes,
    nome_arquivo: str,
    coluna_identificador: str = "matricula",
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Processa um CSV/XLSX e aplica `update` em Servidores existentes.

    NÃO cria servidor novo — o caso de uso é enriquecer cadastros que já
    vieram do legado mas estão com lacunas pré-eSocial. Criação fica para
    o fluxo de admissão manual + admissão em massa do Bloco 4.

    Erros previsíveis (servidor não encontrado, coluna identificador
    ausente) vão pro array `erros` sem abortar — operador vê o resumo.
    """
    if coluna_identificador not in ("matricula", "cpf"):
        raise ValueError("coluna_identificador deve ser 'matricula' ou 'cpf'.")

    try:
        cabecalho, dados = _parsear_arquivo(conteudo_bytes, nome_arquivo)
    except ValueError as exc:
        return {
            "total_linhas": 0,
            "atualizados": 0,
            "ignorados_servidor_nao_encontrado": 0,
            "ignorados_sem_mudanca": 0,
            "erros": [{"linha": 0, "mensagem": str(exc)}],
            "preview": [],
            "colunas_aceitas_mapeadas": [],
            "colunas_ignoradas": [],
        }

    mapeamento = _mapear_colunas(cabecalho)
    colunas_ignoradas = [c for c in cabecalho if c not in mapeamento and c]

    # Localiza a coluna identificadora no cabeçalho real (case insensitive).
    coluna_id_real: str | None = None
    for col in cabecalho:
        chave = _normalizar_cabecalho(col)
        if coluna_identificador == "matricula" and chave in ("matricula", "matrícula"):
            coluna_id_real = col
            break
        if coluna_identificador == "cpf" and chave == "cpf":
            coluna_id_real = col
            break
    if coluna_id_real is None:
        return {
            "total_linhas": len(dados),
            "atualizados": 0,
            "ignorados_servidor_nao_encontrado": 0,
            "ignorados_sem_mudanca": 0,
            "erros": [
                {
                    "linha": 0,
                    "mensagem": f"Coluna identificadora '{coluna_identificador}' não encontrada no cabeçalho.",
                }
            ],
            "preview": [],
            "colunas_aceitas_mapeadas": list(mapeamento.values()),
            "colunas_ignoradas": colunas_ignoradas,
        }

    resultados: list[_LinhaResultado] = []
    for idx, linha in enumerate(dados, start=2):  # linha 1 = cabeçalho
        identificador_bruto = (linha.get(coluna_id_real) or "").strip()
        if not identificador_bruto:
            resultados.append(
                _LinhaResultado(
                    linha=idx,
                    identificador="",
                    servidor_id=None,
                    mensagem_erro="Identificador vazio.",
                )
            )
            continue
        servidor = _resolver_servidor(identificador_bruto, coluna_identificador)
        if servidor is None:
            resultados.append(
                _LinhaResultado(
                    linha=idx,
                    identificador=identificador_bruto,
                    servidor_id=None,
                    mensagem_erro="Servidor não encontrado.",
                )
            )
            continue
        antes: dict[str, Any] = {}
        depois: dict[str, Any] = {}
        for col_origem, campo_modelo in mapeamento.items():
            valor_bruto = linha.get(col_origem, "")
            valor_norm = _normalizar_valor(campo_modelo, valor_bruto)
            if not valor_norm:
                # Não sobrescreve campo preenchido com string vazia.
                continue
            atual = getattr(servidor, campo_modelo) or ""
            if str(atual).strip() == valor_norm:
                continue
            antes[campo_modelo] = atual
            depois[campo_modelo] = valor_norm
        resultados.append(
            _LinhaResultado(
                linha=idx,
                identificador=identificador_bruto,
                servidor_id=servidor.id,
                antes=antes,
                depois=depois,
            )
        )

    if not dry_run:
        _aplicar_mudancas(resultados)

    atualizados = sum(1 for r in resultados if r.servidor_id is not None and r.depois)
    nao_encontrados = sum(
        1 for r in resultados if r.servidor_id is None and r.mensagem_erro == "Servidor não encontrado."
    )
    sem_mudanca = sum(
        1 for r in resultados if r.servidor_id is not None and not r.depois
    )
    erros = [
        {"linha": r.linha, "mensagem": r.mensagem_erro}
        for r in resultados
        if r.mensagem_erro and r.mensagem_erro != "Servidor não encontrado."
    ]
    preview = [
        {
            "linha": r.linha,
            "identificador": r.identificador,
            "servidor_id": r.servidor_id,
            "antes": r.antes,
            "depois": r.depois,
        }
        for r in resultados
        if r.depois
    ][:50]
    return {
        "total_linhas": len(dados),
        "atualizados": atualizados,
        "ignorados_servidor_nao_encontrado": nao_encontrados,
        "ignorados_sem_mudanca": sem_mudanca,
        "erros": erros,
        "preview": preview,
        "colunas_aceitas_mapeadas": sorted(set(mapeamento.values())),
        "colunas_ignoradas": colunas_ignoradas,
        "dry_run": dry_run,
    }


@transaction.atomic
def _aplicar_mudancas(resultados: list[_LinhaResultado]) -> None:
    """Persiste somente linhas com `depois` não-vazio (preserva simple-history)."""
    ids_para_aplicar = {r.servidor_id: r.depois for r in resultados if r.servidor_id and r.depois}
    if not ids_para_aplicar:
        return
    servidores = {s.id: s for s in Servidor.objects.filter(id__in=ids_para_aplicar.keys())}
    for sid, mudancas in ids_para_aplicar.items():
        servidor = servidores.get(sid)
        if not servidor:
            continue
        for campo, valor in mudancas.items():
            setattr(servidor, campo, valor)
        servidor.save()
